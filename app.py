import time
import streamlit as st
import pandas as pd
import re
from tempfile import NamedTemporaryFile
import argostranslate.translate
from io import BytesIO
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl

def download_argos_language_pack():
    from_code = "zh"
    to_code = "en"

    # Download and install Argos Translate package
    argostranslate.package.update_package_index()
    available_packages = argostranslate.package.get_available_packages()
    package_to_install = next(
        filter(
            lambda x: x.from_code == from_code and x.to_code == to_code, available_packages
        )
    )
    argostranslate.package.install_from_path(package_to_install.download())

def has_chinese_chars(text):
    """Check if text contains Chinese characters"""
    return bool(re.search('[\u4e00-\u9fff]', text))

def translate(s):
    if not isinstance(s, str) or not has_chinese_chars(s) or not s:
        return s

    translatedText = argostranslate.translate.translate(s, "zh", "en")

    return translatedText

def load_workbook(uploaded_file):
    """Load workbook with all formatting intact"""
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(uploaded_file.getvalue())
        return openpyxl.load_workbook(tmp.name)

def translate_work_book(wb):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Iterate through all cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value:
                        cell.value = translate(str(cell.value))

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Find the max length of content in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add some padding and set column width
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

def translate_and_save_workbook(wb):
    """Convert workbook back to bytes"""
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        translate_work_book(wb)
        wb.save(tmp.name)
        with open(tmp.name, 'rb') as f:
            return BytesIO(f.read())

def translate_df(df):
    rows = []
    for i in range(len(df)):
        row = df.iloc[i]
        cols = []
        for cell in row:
            result = translate(cell)
            cols.append(result)
        
        rows.append(cols)

    new_df = pd.DataFrame(rows).fillna("")
    return new_df

def translate_file(uploaded_file):
    if uploaded_file is None: return
    file_name = uploaded_file.name

    start_time = time.time()
    with st.spinner('Translating Files...'):
        if file_name.endswith('.xlsx'):
            # Load workbook (preserves all formatting)
            wb = load_workbook(uploaded_file)
            
            # Show sheet info (optional)
            sheet_names = wb.sheetnames
            # st.success(f"Loaded workbook with {len(sheet_names)} sheets: {', '.join(sheet_names)}")
            
            # Download button
            output = translate_and_save_workbook(wb)
            st.download_button(
                label="Download Translated Excel (With Formatting)",
                data=output.getvalue(),
                file_name="translated_" + uploaded_file.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif file_name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
            translated_df = translate_df(df)

            download_file_name = "translated_" + file_name
            csv = translated_df.to_csv(index=False).encode('utf-8')

            download2 = st.download_button(
                label="Download Translated csv",
                data=csv,
                file_name=download_file_name,
                mime='text/csv'
            )
        else:
            st.warning("File must be .csv or .xlsx!")
            return

        end_time = time.time()
        seconds_elapsed = round((end_time - start_time), 2)
        st.success("File translated successfully in " + str(seconds_elapsed) + " seconds")
 
def show_ui():
    # uploaded_file = st.file_uploader("Choose an Excel/csv file", type={"xlsx", 'csv'})
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx', 'csv'])

    if uploaded_file is not None:
        try:
            translate_file(uploaded_file)
        except Exception as e:
            st.error("Translation/Upload failed: " + repr(e))

st.title("Excel and CSV Translator (Chinese to English)")
download_argos_language_pack()
show_ui()
